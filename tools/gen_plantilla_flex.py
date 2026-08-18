"""Genera la plantilla Excel para importar datos léxicos a FLEx.

    python tools/gen_plantilla_flex.py ame          # las dos ortografías
    python tools/gen_plantilla_flex.py ame --sil    # solo el alfabeto SIL
    python tools/gen_plantilla_flex.py ame --todo   # las 1681 entradas

Sale en templates/flex/Plantilla-FLEx-<Lengua>.xlsx. Las filas de ejemplo se
sacan del diccionario ya construido del proyecto (out/*.tsv), así el equipo ve
sus propias palabras en la estructura que FLEx espera.

La fila 1 lleva el código de columna (= marcador SFM) y la fila 2 la etiqueta
en español. Los datos empiezan en la fila 3. No borrar la fila 2: el conversor
la salta siempre.
"""

import csv
import json
import sys
import unicodedata as ud
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

RAIZ = Path(__file__).resolve().parent.parent

# código, etiqueta, obligatorio, ancho, ayuda (va como comentario en la celda)
COLUMNAS = [
    ("lx", "Lexema (ortografía SIL)", True, 24,
     "Forma de la entrada. FLEx la usa como marcador de registro: SIEMPRE la "
     "primera columna y nunca vacía.\n\nAfijos con guion: pe- (prefijo), -et "
     "(sufijo). FLEx deduce el tipo de morfema del guion."),
    ("lxo", "Lexema (alfabeto oficial MINEDU)", False, 24,
     "La misma forma en el alfabeto oficial (digrafos en vez de diacríticos). "
     "Se importa como un segundo sistema de escritura vernáculo.\n\nDéjala "
     "vacía si todavía no se hizo la conversión ortográfica."),
    ("hm", "N.º de homónimo", False, 9,
     "1, 2, 3... solo cuando dos entradas distintas comparten la misma forma. "
     "Si no hay homónimos, déjala vacía."),
    ("lc", "Forma de citación", False, 20,
     "Solo si la forma que se muestra en el diccionario difiere del lexema "
     "(p. ej. el lexema es una raíz y se cita la forma flexionada)."),
    ("va", "Variante", False, 18,
     "Otra forma de la misma palabra (dialectal, libre). Una por fila; si hay "
     "varias, repite la entrada."),
    ("ps", "Categoría gramatical", True, 12,
     "Usa las abreviaturas de la hoja «Categorías», siempre iguales. Cada "
     "variante de escritura crea una categoría duplicada en FLEx."),
    ("sn", "N.º de acepción", False, 9,
     "1, 2, 3... Una fila POR ACEPCIÓN. Repite el lexema en la columna lx en "
     "cada fila de la misma entrada.\n\nSi la entrada tiene una sola acepción, "
     "deja esta celda vacía."),
    ("ge", "Glosa (español)", True, 30,
     "Traducción corta, una o dos palabras. Es lo que aparece en el "
     "interlineal. NO pongas varias glosas separadas por ';' — cada una va en "
     "su propia fila como acepción aparte."),
    ("de", "Definición (español)", False, 40,
     "Definición en oración completa. Opcional pero muy recomendable."),
    ("sd", "Dominio semántico", False, 16,
     "Código de la lista SIL, p. ej. 1.5.1 o 5.2. Ver la hoja «Dominios»."),
    ("rf", "Referencia del ejemplo", False, 14,
     "De dónde salió el ejemplo: 1SA 1:10, entrevista 2026-03, etc."),
    ("xv", "Ejemplo (yanesha')", False, 42, "Oración de ejemplo en la lengua."),
    ("xn", "Traducción del ejemplo", False, 42,
     "La misma oración en español."),
    ("re", "Entrada reversa (español)", False, 20,
     "Palabra española bajo la cual debe aparecer esta entrada en el índice "
     "español→yanesha'. Si está vacía, FLEx usa la glosa."),
    ("cf", "Véase también", False, 18,
     "Lexema de otra entrada relacionada. Debe existir como \\lx en alguna "
     "fila, si no FLEx crea una entrada fantasma."),
    ("sy", "Sinónimo", False, 18, "Lexema de una entrada sinónima."),
    ("an", "Antónimo", False, 18, "Lexema de una entrada antónima."),
    ("et", "Etimología", False, 22, "Forma de origen (préstamo, protoforma)."),
    ("bw", "Préstamo de", False, 14,
     "Lengua de origen si es préstamo: español, quechua, ashéninka..."),
    ("nt", "Nota", False, 30, "Nota general para el equipo."),
    ("st", "Confianza", False, 10,
     "A = confirmado por el equipo. B = propuesto, falta revisar. C = dudoso.\n\n"
     "No es un campo MDF: en FLEx se mapea a Sense > Status o se descarta."),
    ("lem", "Lema hebreo/griego", False, 18,
     "Para proyectos de traducción bíblica: el lema del texto fuente que esta "
     "entrada traduce. Requiere un campo personalizado en FLEx."),
]

CATEGORIAS = [
    ("s", "Sustantivo", "Noun"),
    ("v", "Verbo", "Verb"),
    ("vt", "Verbo transitivo", "Transitive Verb"),
    ("vi", "Verbo intransitivo", "Intransitive Verb"),
    ("adj", "Adjetivo", "Adjective"),
    ("adv", "Adverbio", "Adverb"),
    ("pron", "Pronombre", "Pronoun"),
    ("dem", "Demostrativo", "Demonstrative"),
    ("num", "Numeral", "Numeral"),
    ("conj", "Conjunción", "Conjunction"),
    ("interj", "Interjección", "Interjection"),
    ("post", "Posposición", "Postposition"),
    ("clas", "Clasificador", "Classifier"),
    ("pref", "Prefijo", "Prefix"),
    ("suf", "Sufijo", "Suffix"),
    ("nprop", "Nombre propio", "Proper Noun"),
    ("fras", "Frase / locución", "(Frase: FLEx la marca sola por el espacio)"),
]

DOMINIOS = [
    ("1", "Universo, creación", "cielo, tierra, animales, plantas, clima"),
    ("2", "Persona", "cuerpo, salud, nacer, morir, edades"),
    ("3", "Lengua y pensamiento", "hablar, saber, creer, emociones, voluntad"),
    ("4", "Comportamiento social", "familia, autoridad, ley, religión, guerra"),
    ("5", "Vida diaria", "casa, comida, ropa, higiene"),
    ("6", "Trabajo y ocupaciones", "agricultura, caza, pesca, oficios, comercio"),
    ("7", "Acciones físicas", "mover, llevar, poner, golpear, agarrar"),
    ("8", "Estados", "cantidad, tamaño, forma, color, tiempo, calidad"),
    ("9", "Gramática", "afijos, pronombres, conectores, partículas"),
]

INSTRUCCIONES = [
    ("Cómo llenar esta plantilla y llevarla a FLEx", None),
    ("", None),
    ("Antes de empezar", None),
    ("1", "Haz una copia de respaldo del proyecto FLEx (File > Back up this "
          "Project). Una importación mal mapeada es difícil de deshacer."),
    ("2", "En FLEx crea una entrada de prueba a mano con TODOS los campos que "
          "vas a usar, y expórtala como SFM (File > Export > Standard Format). "
          "Ese archivo te muestra los marcadores exactos que tu proyecto "
          "espera. Es el truco más útil de toda la lista de FLEx."),
    ("", None),
    ("Llenando la hoja «Léxico»", None),
    ("3", "Una fila por ACEPCIÓN, no por entrada. Si una palabra tiene tres "
          "significados, son tres filas: repite el lexema en la columna lx y "
          "numera 1, 2, 3 en la columna sn."),
    ("4", "Nunca metas varias glosas en una sola celda separadas por ';'. Eso "
          "entra a FLEx como una glosa larga y absurda."),
    ("5", "La columna lx no puede quedar vacía en ninguna fila: FLEx la usa "
          "como marcador de registro."),
    ("6", "Escribe las categorías gramaticales SIEMPRE igual. 'v' y 'V' crean "
          "dos categorías distintas en FLEx, y fusionarlas después es un lío."),
    ("7", "No uses Buscar y reemplazar de Excel sobre la columna lx sin "
          "revisar: rompe los diacríticos compuestos (c̈, p̃, t̃)."),
    ("", None),
    ("Convirtiendo a SFM", None),
    ("8", "Guarda este archivo y ejecuta en la terminal:"),
    ("", "    python templates/flex/excel_a_flex.py \"ruta/al/archivo.xlsx\""),
    ("9", "Eso genera un .sfm y un .lift junto al Excel. El conversor ya "
          "resuelve los dos problemas clásicos de SheetSwiper: agrupa bien las "
          "acepciones múltiples y no deja bloques de acepción vacíos."),
    ("", "(SheetSwiper también sirve, pero exige .xls antiguo y produce "
         "bloques vacíos que hay que limpiar con regex antes de importar.)"),
    ("", None),
    ("Importando en FLEx", None),
    ("10", "LIFT (recomendado): File > Import > Lexicon (LIFT). Conserva mejor "
           "la estructura de acepciones y no pide mapear nada a mano."),
    ("11", "SFM: File > Import > Standard Format Marker Data. En la ventana de "
           "mapeo, revisa marcador por marcador."),
    ("12", "Los marcadores que caigan en «Residue» no se importan. Selecciona "
           "cada uno, pulsa «Modify» y dile a FLEx a qué campo va. Es normal "
           "que \\hm, \\et y \\va caigan ahí la primera vez."),
    ("13", "Para \\lxo (alfabeto oficial): mapéalo también a Lexeme Form, pero "
           "eligiendo el segundo sistema de escritura vernáculo. Tienes que "
           "haberlo creado antes en Format > Set up Vernacular Writing Systems."),
    ("14", "Para \\st y \\lem necesitas campos personalizados (Tools > "
           "Configure > Custom Fields). Si no los quieres, déjalos en Residue."),
    ("", None),
    ("Después de importar", None),
    ("15", "Revisa Lexicon > Bulk Edit Entries y ordena por categoría "
           "gramatical: si aparecen categorías duplicadas, se colaron por "
           "escritura inconsistente en la columna ps."),
    ("16", "Si el orden alfabético sale raro, es el ordenamiento de la lengua: "
           "hay que definir reglas ICU personalizadas para el alfabeto yanesha'."),
]

FONDO_OBLIG = PatternFill("solid", fgColor="1F4E5F")
FONDO_OPC = PatternFill("solid", fgColor="4A5A63")
FONDO_ETIQ = PatternFill("solid", fgColor="E8EDEF")
FONDO_TIT = PatternFill("solid", fgColor="1F4E5F")
BORDE = Border(bottom=Side("thin", color="B0BEC5"))


def _titulo(ws, texto, ancho):
    ws.append([texto])
    c = ws.cell(ws.max_row, 1)
    c.font = Font(bold=True, size=13, color="FFFFFF")
    c.fill = FONDO_TIT
    ws.merge_cells(start_row=ws.max_row, start_column=1,
                   end_row=ws.max_row, end_column=ancho)
    ws.append([])


def hoja_lexico(wb, filas, columnas):
    ws = wb.create_sheet("Léxico")

    for i, (cod, etiq, oblig, ancho, ayuda) in enumerate(columnas, start=1):
        letra = get_column_letter(i)
        ws.column_dimensions[letra].width = ancho

        c = ws.cell(1, i, cod)
        c.font = Font(bold=True, size=11, color="FFFFFF", name="Consolas")
        c.fill = FONDO_OBLIG if oblig else FONDO_OPC
        c.alignment = Alignment(horizontal="center")
        c.comment = Comment(f"{etiq}\n\n{ayuda}", "kamus-toolkit", height=200,
                            width=340)

        e = ws.cell(2, i, etiq + (" *" if oblig else ""))
        e.font = Font(italic=True, size=9, color="37474F")
        e.fill = FONDO_ETIQ
        e.alignment = Alignment(wrap_text=True, vertical="top")
        e.border = BORDE

    ws.row_dimensions[2].height = 30

    codigos = [c[0] for c in columnas]
    for fila in filas:
        ws.append([fila.get(c, "") for c in codigos])

    # Las columnas en lengua vernácula necesitan una fuente que dibuje bien
    # los diacríticos combinantes (c̈, p̃, t̃).
    vernaculas = [codigos.index(c) + 1 for c in ("lx", "lxo", "xv")
                  if c in codigos]
    for r in range(3, ws.max_row + 1):
        for col in vernaculas:
            ws.cell(r, col).font = Font(name="Charis SIL", size=11)

    ws.freeze_panes = "C3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(columnas))}{ws.max_row}"

    tope = max(ws.max_row + 400, 500)
    dv_ps = DataValidation(
        type="list", allow_blank=True,
        formula1='"' + ",".join(c[0] for c in CATEGORIAS) + '"',
        error="Usa una abreviatura de la hoja «Categorías».",
        errorTitle="Categoría no reconocida", showErrorMessage=False)
    ws.add_data_validation(dv_ps)
    col_ps = get_column_letter(codigos.index("ps") + 1)
    dv_ps.add(f"{col_ps}3:{col_ps}{tope}")

    dv_st = DataValidation(
        type="list", allow_blank=True, formula1='"A,B,C"',
        showErrorMessage=False)
    ws.add_data_validation(dv_st)
    col_st = get_column_letter(codigos.index("st") + 1)
    dv_st.add(f"{col_st}3:{col_st}{tope}")

    return ws


def hoja_campos(wb, columnas):
    ws = wb.create_sheet("Campos")
    _titulo(ws, "Qué es cada columna", 5)
    ws.append(["Columna", "Marcador SFM", "Campo en FLEx", "¿Obligatorio?",
               "Descripción"])
    for c in ws[ws.max_row]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = FONDO_OBLIG

    destinos = {
        "lx": "Lexeme Form (sistema vernáculo 1)",
        "lxo": "Lexeme Form (sistema vernáculo 2) — requiere Modify",
        "hm": "Homograph Number — suele caer en Residue",
        "lc": "Citation Form",
        "va": "Variant Form — suele caer en Residue",
        "ps": "Sense > Grammatical Info",
        "sn": "(estructural: separa las acepciones)",
        "ge": "Sense > Gloss",
        "de": "Sense > Definition",
        "sd": "Sense > Semantic Domain",
        "rf": "Example > Reference",
        "xv": "Example Sentence",
        "xn": "Example > Translation",
        "re": "Sense > Reversal Entry",
        "cf": "Lexical Relation: See also",
        "sy": "Lexical Relation: Synonym",
        "an": "Lexical Relation: Antonym",
        "et": "Etymology > Form — suele caer en Residue",
        "bw": "Etymology > Source language",
        "nt": "Note (general)",
        "st": "Campo personalizado o Sense > Status",
        "lem": "Campo personalizado (no es MDF)",
    }
    for cod, etiq, oblig, _, ayuda in columnas:
        ws.append([cod, "\\" + cod, destinos.get(cod, ""),
                   "sí" if oblig else "", f"{etiq}. {ayuda.splitlines()[0]}"])

    for w, letra in zip((11, 13, 44, 13, 88), "ABCDE"):
        ws.column_dimensions[letra].width = w
    for r in ws.iter_rows(min_row=4):
        r[4].alignment = Alignment(wrap_text=True, vertical="top")
        r[0].font = Font(name="Consolas")
        r[1].font = Font(name="Consolas")
    return ws


def hoja_categorias(wb):
    ws = wb.create_sheet("Categorías")
    _titulo(ws, "Abreviaturas para la columna «ps»", 3)
    ws.append(["Escribe esto", "Significa", "Categoría de FLEx a la que mapear"])
    for c in ws[ws.max_row]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = FONDO_OBLIG
    for fila in CATEGORIAS:
        ws.append(list(fila))
    for w, letra in zip((14, 26, 46), "ABC"):
        ws.column_dimensions[letra].width = w

    ws.append([])
    ws.append(["Esta lista es un punto de partida. Ajústenla al análisis "
               "gramatical del yanesha' — lo único que importa es que dentro "
               "del equipo se escriba siempre igual."])
    ws.cell(ws.max_row, 1).font = Font(italic=True, size=9, color="546E7A")
    ws.merge_cells(start_row=ws.max_row, start_column=1,
                   end_row=ws.max_row, end_column=3)
    ws.cell(ws.max_row, 1).alignment = Alignment(wrap_text=True)
    return ws


def hoja_dominios(wb):
    ws = wb.create_sheet("Dominios")
    _titulo(ws, "Dominios semánticos SIL (primer nivel)", 3)
    ws.append(["Código", "Dominio", "Incluye"])
    for c in ws[ws.max_row]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = FONDO_OBLIG
    for fila in DOMINIOS:
        ws.append(list(fila))
    for w, letra in zip((10, 30, 56), "ABC"):
        ws.column_dimensions[letra].width = w

    ws.append([])
    ws.append(["Los subdominios (1.5.1, 2.3.4...) están en la lista completa de "
               "FLEx: Lists > Semantic Domains. Basta con el primer nivel para "
               "empezar; se puede refinar después dentro de FLEx."])
    ws.cell(ws.max_row, 1).font = Font(italic=True, size=9, color="546E7A")
    ws.merge_cells(start_row=ws.max_row, start_column=1,
                   end_row=ws.max_row, end_column=3)
    ws.cell(ws.max_row, 1).alignment = Alignment(wrap_text=True)
    return ws


def hoja_instrucciones(wb, lengua, instrucciones):
    ws = wb.create_sheet("Instrucciones", 0)
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 104

    ws.append([f"Plantilla de importación léxica a FLEx — {lengua}"])
    c = ws.cell(1, 1)
    c.font = Font(bold=True, size=14, color="FFFFFF")
    c.fill = FONDO_TIT
    ws.merge_cells("A1:B1")
    ws.row_dimensions[1].height = 24
    ws.append([])

    for num, texto in instrucciones:
        if texto is None:
            if not num:
                ws.append([])
                continue
            ws.append([num])
            cc = ws.cell(ws.max_row, 1)
            cc.font = Font(bold=True, size=11, color="1F4E5F")
            ws.merge_cells(start_row=ws.max_row, start_column=1,
                           end_row=ws.max_row, end_column=2)
            continue
        ws.append([num, texto])
        ws.cell(ws.max_row, 1).font = Font(bold=True, color="1F4E5F")
        ws.cell(ws.max_row, 1).alignment = Alignment(horizontal="right",
                                                     vertical="top")
        cc = ws.cell(ws.max_row, 2)
        cc.alignment = Alignment(wrap_text=True, vertical="top")
        if texto.startswith("    "):
            cc.font = Font(name="Consolas", size=10)
    return ws


def _anotar(filas):
    """Datos que el diccionario automático no puede deducir solo.

    Definiciones, dominios y el ejemplo del corpus. Se aplican igual a la
    plantilla de muestra que a la exportación completa.
    """
    por_glosa = {(f["lx"], f["ge"]): f for f in filas}

    f = por_glosa.get(("acheñ", "persona"))
    if f:
        f["de"] = "Ser humano, individuo de la comunidad."
        f["sd"] = "2"
        f["re"] = "persona"
        f["cf"] = "acheñeneshá"

    f = por_glosa.get(("acheñochayayo", "humanidad"))
    if f:
        f["de"] = "El conjunto de los seres humanos."
        f["sd"] = "2"

    f = por_glosa.get(("aarón", "de Aarón"))
    if f:
        f["bw"] = "hebreo"
        f["et"] = "אַהֲרֹן"
        f["nt"] = "Nombre propio: revisar con el equipo la forma de citación."

    # Ejemplo tomado del corpus alineado del propio proyecto: 1 Samuel 1:10,
    # la oración de Ana. Todas sus palabras existen ya como entradas del
    # diccionario, así que el ejemplo es coherente con el resto de la hoja.
    oracion = ("Ana éñe patsrrémueñot̃ póchtsohuen máyochana parets "
               "ñam̃a yáhua t̃ematen")
    traduccion = ("Ella, con amargura de todo su ser, oró a Jehová y lloró "
                  "abundantemente")

    f = por_glosa.get(("máyochana", "oraba"))
    if f:
        f["de"] = "Dirigir la palabra a Dios."
        f["sd"] = "4"
        f["rf"] = "1SA 1:10"
        f["xv"] = oracion
        f["xn"] = traduccion

    f = por_glosa.get(("póchtsohuen", "alma"))
    if f:
        f["de"] = "El interior de la persona, su ser entero."
        f["sd"] = "2"
        f["re"] = "alma"

    f = por_glosa.get(("parets", "Jehová"))
    if f:
        f["nt"] = "Término clave: un cambio aquí afecta a todo el corpus."



def _leer_diccionario(proyecto):
    """Lee el TSV del diccionario ya construido. Devuelve (lista, indice)."""
    ruta = RAIZ / "projects" / proyecto / "project.json"
    cfg = json.loads(ruta.read_text(encoding="utf-8"))
    tsv = RAIZ / "projects" / proyecto / "out" / (cfg["output"]["basename"] + ".tsv")

    orden, dic = [], {}
    with tsv.open(encoding="utf-8", newline="") as fh:
        for v in csv.DictReader(fh, delimiter="	"):
            v["origen"] = ud.normalize("NFC", v["origen"])
            orden.append(v)
            dic[v["origen"]] = v
    return orden, dic


def _afijos(proyecto):
    """Formas que el equipo segmentó como afijos, con su dirección.

    Devuelve forma-sin-guion -> ('pref'|'suf'|None, ejemplo). None cuando la
    misma forma aparece como prefijo y como sufijo: esa dirección la decide
    el equipo, no este script.
    """
    morf = json.loads((RAIZ / "projects" / proyecto / "morphology.json")
                      .read_text(encoding="utf-8"))
    acc = {}
    for afijo, funcion, ejemplo in morf:
        forma = ud.normalize("NFC", afijo.strip("-"))
        tipo = "pref" if afijo.endswith("-") else "suf"
        if forma in acc and acc[forma][0] != tipo:
            acc[forma] = (None, acc[forma][1])
        else:
            acc[forma] = (tipo, ejemplo)
    return acc


def muestras_del_proyecto(proyecto):
    """Filas de ejemplo, una por acepción, del diccionario ya construido."""
    _, dic = _leer_diccionario(proyecto)

    # Entradas elegidas a mano: cubren acepciones múltiples, frases, afijos,
    # préstamos y los diacríticos compuestos c̈ p̃ t̃ m̃.
    elegidas = [
        ("acheñ", "s", None),
        ("acheret̃", "adj/s", None),
        ("acheñochayayo", "s", None),
        ("ac̈horene aser", "fras", None),
        ("áqueshp̃at", "v", None),
        ("ácohuentataret̃terra", "adj", None),
        ("ácohuentathuan", "v", None),
        ("ác̈homenó", "adj", None),
        ("áchasar", "v", None),
        ("acraret̃esho", "s", None),
        ("aj", "s", None),
        ("ahuamencat̃", "adj", None),
        ("achtaret̃ puechemeretam̃", "fras", None),
        ("aarón", "nprop", None),
        # Las tres siguientes salen de la oración de ejemplo de más abajo,
        # para que el ejemplo quede anclado a entradas que existen de verdad.
        ("máyochana", "v", None),
        ("póchtsohuen", "s", None),
        ("parets", "nprop", None),
    ]
    # 'acheret̃' cambia de categoría entre acepciones: adj para 'asado',
    # s para 'pan de la Presencia'. Se resuelve fila por fila más abajo.
    ps_por_glosa = {
        "asado": "adj", "yuca asada": "s", "pan de la proposición": "s",
        "pan de la Presencia": "s", "pan sagrado": "s",
    }

    filas = []
    for forma, ps, _ in elegidas:
        v = dic.get(ud.normalize("NFC", forma))
        if v is None:
            print(f"  aviso: '{forma}' no está en el diccionario, se omite")
            continue
        glosas = [g.strip() for g in v["glosa"].split(";") if g.strip()]
        multi = len(glosas) > 1
        for i, glosa in enumerate(glosas, start=1):
            fila = {
                "lx": forma,
                "ps": ps_por_glosa.get(glosa, ps if "/" not in ps else "s"),
                "sn": i if multi else "",
                "ge": glosa,
                "st": v["tingkat"],
                "lem": v["lema_yunani_ibrani"],
            }
            filas.append(fila)

    _anotar(filas)

    # Dos afijos, para mostrar la convención del guion.
    morf = json.loads((RAIZ / "projects" / proyecto / "morphology.json")
                      .read_text(encoding="utf-8"))
    for afijo, funcion, ejemplo in morf:
        if afijo in ("pe-", "-et"):
            filas.append({
                "lx": afijo,
                "ps": "pref" if afijo.endswith("-") else "suf",
                "ge": funcion.split(";")[0].strip(),
                "nt": f"Segmentación del equipo en Paratext: {ejemplo}",
                "st": "B",
            })

    return filas


def variante_sil(columnas, instrucciones):
    """Versión para equipos que trabajan solo en el alfabeto SIL.

    Quita la columna del alfabeto oficial MINEDU y todo lo que hable de un
    segundo sistema de escritura. Lo demás —normalización NFC, fuentes para
    los diacríticos combinantes, estructura de acepciones— se mantiene igual,
    porque no depende de qué alfabeto se use.
    """
    cols = []
    for cod, etiq, oblig, ancho, ayuda in columnas:
        if cod == "lxo":
            continue
        if cod == "lx":
            etiq = "Lexema (alfabeto SIL)"
        cols.append((cod, etiq, oblig, ancho, ayuda))

    instr, n = [], 0
    for num, texto in instrucciones:
        if texto and ("lxo" in texto or "MINEDU" in texto):
            continue
        if texto is not None and num:
            n += 1
            num = str(n)
        instr.append((num, texto))
    instr.append(("", None))
    instr.append(("Sobre el alfabeto", None))
    instr.append((str(n + 1),
                  "Esta versión cubre solo el alfabeto SIL, el de los "
                  "diacríticos (c̈, p̃, t̃, m̃). Si algún día hay que pasar al "
                  "alfabeto oficial del MINEDU, que usa digrafos, la "
                  "herramienta es SIL Converters con una tabla TECkit de "
                  "correspondencias; no se hace a mano ni con Buscar y "
                  "reemplazar."))
    return cols, instr


def todas_las_entradas(proyecto, prop=None):
    """Las 1681 entradas del diccionario, una fila por acepcion.

    Rellena lo que se puede deducir con certeza del propio corpus y deja en
    blanco lo que no. La categoria gramatical es justamente lo que el TSV no
    tiene: solo se puede fijar en tres casos -- frases, afijos y nombres
    propios. El resto lo pone el equipo, que es quien sabe.
    """
    entradas, _ = _leer_diccionario(proyecto)
    afijos = _afijos(proyecto)

    filas = []
    for v in entradas:
        forma = v["origen"]
        glosas = [g.strip() for g in v["glosa"].split(";") if g.strip()]
        multi = len(glosas) > 1

        ps = ""
        nota = ""
        if v["jenis"] == "frase":
            # FLEx marca como frase cualquier lexema con espacio, asi que aqui
            # solo estamos adelantando lo que haria de todos modos.
            ps = "fras"
        elif forma in afijos:
            tipo, ejemplo = afijos[forma]
            nota = f"Afijo segmentado por el equipo en Paratext: {ejemplo}"
            if tipo:
                ps = tipo
            else:
                nota += (". Aparece como prefijo y como sufijo: el equipo debe "
                         "decidir cual es cual antes de importar.")
            nota += (". FLEx necesita el guion en la forma (pe-, -et) para "
                     "asignar el tipo de morfema.")

        for i, glosa in enumerate(glosas, start=1):
            ps_fila = ps
            if not ps_fila and prop and _es_nombre_propio(glosa, prop):
                ps_fila = "nprop"
            fila = {
                "lx": forma,
                "ps": ps_fila,
                "sn": i if multi else "",
                "ge": glosa,
                "st": v["tingkat"],
                "lem": v["lema_yunani_ibrani"],
            }
            if nota:
                fila["nt"] = nota
            filas.append(fila)

    _anotar(filas)
    return filas


def _es_nombre_propio(glosa, prop):
    """La glosa es un nombre propio, quiza con marca de caso ('de Aaron')."""
    if glosa in prop:
        return True
    return any(t in prop for t in glosa.replace(",", " ").split()
               if t[:1].isupper())


def _cargar_prop(proyecto):
    """Lista de nombres propios del proyecto, si existe."""
    ruta = RAIZ / "projects" / proyecto / "prop_map.py"
    if not ruta.exists():
        return None
    espacio = {}
    exec(compile(ruta.read_text(encoding="utf-8"), str(ruta), "exec"), espacio)
    return espacio.get("PROP")


def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    sil = "--sil" in sys.argv
    todo = "--todo" in sys.argv

    proyecto = args[0] if args else "ame"
    cfg = json.loads((RAIZ / "projects" / proyecto / "project.json")
                     .read_text(encoding="utf-8"))
    lengua = cfg["language"]["name"]

    columnas, instrucciones = COLUMNAS, INSTRUCCIONES
    sufijo = ""
    if sil:
        columnas, instrucciones = variante_sil(COLUMNAS, INSTRUCCIONES)
        sufijo = "-SIL"

    if todo:
        filas = todas_las_entradas(proyecto, _cargar_prop(proyecto))
        base = "Lexico-FLEx"
        sin_ps = sum(1 for f in filas if not f.get("ps"))
        print(f"  {sin_ps} de {len(filas)} acepciones quedan sin categoria "
              f"gramatical: la columna ps la llena el equipo.")
    else:
        filas = muestras_del_proyecto(proyecto)
        base = "Plantilla-FLEx"

    wb = Workbook()
    wb.remove(wb.active)
    hoja_lexico(wb, filas, columnas)
    hoja_campos(wb, columnas)
    hoja_categorias(wb)
    hoja_dominios(wb)
    hoja_instrucciones(wb, lengua, instrucciones)
    wb.active = 0

    destino = RAIZ / "templates" / "flex"
    destino.mkdir(parents=True, exist_ok=True)
    salida = destino / f"{base}-{lengua.replace(chr(39), '')}{sufijo}.xlsx"
    wb.save(salida)
    clase = "filas" if todo else "filas de ejemplo"
    print(f"{salida}  ({len(filas)} {clase}, {len(columnas)} columnas)")


if __name__ == "__main__":
    main()
