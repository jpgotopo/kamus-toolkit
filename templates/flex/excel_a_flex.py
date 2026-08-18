"""Convierte la plantilla Excel en los archivos que FLEx sabe importar.

    python excel_a_flex.py Plantilla-FLEx-Yanesha-SIL.xlsx

Produce dos archivos junto al Excel:

    <nombre>.sfm    formato SFM/MDF, para File > Import > Standard Format
    <nombre>.lift   formato LIFT,   para File > Import > Lexicon (LIFT)

Usa LIFT si puedes: no hay que mapear marcadores a mano y la estructura de
acepciones llega intacta.

Por qué no SheetSwiper: SheetSwiper trata cada fila como una entrada, así que
una palabra con tres acepciones entra como tres entradas sueltas, y las
columnas vacías dejan bloques \\ps \\sn \\ge huecos que hay que limpiar con
expresiones regulares antes de importar. Este script agrupa las filas que
comparten lexema y no escribe nunca un campo vacío.

Requiere: pip install openpyxl
"""

import re
import sys
import unicodedata as ud
import uuid

from datetime import date
from pathlib import Path
from xml.sax.saxutils import escape

from openpyxl import load_workbook

HOJA = "Léxico"
FILA_DATOS = 3  # fila 1 = códigos, fila 2 = etiquetas en español

# Campos de entrada y de acepción. El orden es el que MDF espera: primero
# los datos de la entrada, luego cada acepción, y al final las notas.
CAMPOS_ENTRADA = ["lxo", "hm", "lc", "va"]
CAMPOS_ACEPCION = ["ps", "sn", "ge", "de", "sd", "rf", "xv", "xn", "re",
                   "cf", "sy", "an"]
CAMPOS_COLA = ["et", "bw", "nt", "st", "lem"]

# Espacio de nombres propio para que los GUID de LIFT sean estables entre
# ejecuciones: reimportar el mismo Excel actualiza las entradas en vez de
# duplicarlas.
NS = uuid.UUID("6f9d2c3e-0b1a-4d5e-9f80-1c2d3e4f5a6b")

VERNACULO = "ame"
VERNACULO_OFICIAL = "ame-x-oficial"
ANALISIS = "es"


class ErrorPlantilla(Exception):
    pass


def leer_filas(ruta):
    wb = load_workbook(ruta, data_only=True)
    if HOJA not in wb.sheetnames:
        raise ErrorPlantilla(
            f"El archivo no tiene una hoja «{HOJA}». Hojas encontradas: "
            + ", ".join(wb.sheetnames))
    ws = wb[HOJA]

    codigos = []
    for celda in ws[1]:
        v = celda.value
        codigos.append(str(v).strip() if v is not None else None)
    if not codigos or codigos[0] != "lx":
        raise ErrorPlantilla(
            "La primera columna de la fila 1 debe ser «lx». FLEx usa el lexema "
            "como marcador de registro y tiene que ir primero.")

    filas = []
    for n, celdas in enumerate(ws.iter_rows(min_row=FILA_DATOS), start=FILA_DATOS):
        fila = {}
        for cod, celda in zip(codigos, celdas):
            if not cod or celda.value is None:
                continue
            texto = str(celda.value).strip()
            if texto:
                # FLEx guarda internamente en NFD, pero un Excel editado en
                # varias máquinas mezcla formas compuestas y descompuestas.
                # Unificamos en NFC para que 'acheñ' no entre dos veces.
                fila[cod] = ud.normalize("NFC", texto)
        if fila:
            fila["_fila"] = n
            filas.append(fila)
    return filas


def agrupar(filas):
    """Agrupa las filas en entradas. Cada fila es una acepción."""
    entradas = []
    indice = {}
    problemas = []

    for fila in filas:
        lx = fila.get("lx")
        if not lx:
            problemas.append(
                f"fila {fila['_fila']}: sin lexema en la columna lx. Repite el "
                f"lexema en cada fila de acepción.")
            continue
        if not fila.get("ge"):
            problemas.append(
                f"fila {fila['_fila']}: «{lx}» no tiene glosa (columna ge).")

        clave = (lx, fila.get("hm", ""))
        if clave not in indice:
            entrada = {"lx": lx, "acepciones": [], "_fila": fila["_fila"]}
            for c in CAMPOS_ENTRADA + CAMPOS_COLA:
                if fila.get(c):
                    entrada[c] = fila[c]
            indice[clave] = entrada
            entradas.append(entrada)
        else:
            entrada = indice[clave]
            # Datos de nivel de entrada repetidos en la 2.ª fila: los tomamos
            # solo si la primera fila los dejó vacíos.
            for c in CAMPOS_ENTRADA + CAMPOS_COLA:
                if fila.get(c) and not entrada.get(c):
                    entrada[c] = fila[c]

        acepcion = {c: fila[c] for c in CAMPOS_ACEPCION if fila.get(c)}
        if acepcion:
            entrada["acepciones"].append(acepcion)

    for e in entradas:
        if not e["acepciones"]:
            problemas.append(
                f"fila {e['_fila']}: «{e['lx']}» quedó sin ninguna acepción.")

    # Aviso agregado, no fila por fila: en una exportacion completa la
    # categoria gramatical suele faltar en casi todo, y listar cada caso
    # solo enterraria los problemas de verdad.
    sin_ps = sum(1 for f in filas if f.get("ge") and not f.get("ps"))
    if sin_ps:
        problemas.append(
            f"{sin_ps} acepciones sin categoria gramatical (columna ps). "
            f"Se importan igual, pero quedan sin Grammatical Info en FLEx.")

    for g in [g for g in (f.get("ge", "") for f in filas) if ";" in g]:
        problemas.append(
            f"la glosa «{g}» lleva punto y coma: separa cada significado en su "
            f"propia fila con su número en la columna sn.")

    return entradas, problemas


def escribir_sfm(entradas, ruta):
    hoy = date.today().strftime("%d/%b/%Y")
    partes = []
    for e in entradas:
        lineas = [f"\\lx {e['lx']}"]
        for c in CAMPOS_ENTRADA:
            if e.get(c):
                lineas.append(f"\\{c} {e[c]}")

        varias = len(e["acepciones"]) > 1
        for i, ac in enumerate(e["acepciones"], start=1):
            if ac.get("ps"):
                lineas.append(f"\\ps {ac['ps']}")
            # MDF omite \sn cuando la entrada tiene una sola acepción.
            if varias:
                lineas.append(f"\\sn {ac.get('sn', i)}")
            for c in ["ge", "de", "sd", "rf", "xv", "xn", "re", "cf", "sy", "an"]:
                if ac.get(c):
                    lineas.append(f"\\{c} {ac[c]}")

        for c in CAMPOS_COLA:
            if e.get(c):
                lineas.append(f"\\{c} {e[c]}")
        lineas.append(f"\\dt {hoy}")
        partes.append("\n".join(lineas))

    # UTF-8 sin BOM y saltos \n: el importador de FLEx se atraganta con los
    # saltos de línea de Mac y con el BOM en la primera entrada.
    ruta.write_text("\n\n".join(partes) + "\n", encoding="utf-8", newline="\n")


def _forma(lang, texto):
    return (f'<form lang="{lang}"><text>{escape(texto)}</text></form>')


def escribir_lift(entradas, ruta):
    hoy = date.today().isoformat() + "T00:00:00Z"
    out = ['<?xml version="1.0" encoding="UTF-8"?>',
           '<lift version="0.13" producer="kamus-toolkit/excel_a_flex">']

    for e in entradas:
        guid = uuid.uuid5(NS, e["lx"] + "|" + e.get("hm", ""))
        out.append(f'  <entry id="{escape(e["lx"])}_{guid}" guid="{guid}" '
                   f'dateCreated="{hoy}">')

        out.append("    <lexical-unit>")
        out.append("      " + _forma(VERNACULO, e["lx"]))
        if e.get("lxo"):
            out.append("      " + _forma(VERNACULO_OFICIAL, e["lxo"]))
        out.append("    </lexical-unit>")

        if e.get("lc"):
            out.append(f'    <citation>{_forma(VERNACULO, e["lc"])}</citation>')
        if e.get("va"):
            out.append(f'    <variant>{_forma(VERNACULO, e["va"])}</variant>')
        if e.get("et"):
            out.append('    <etymology type="proto" '
                       f'source="{escape(e.get("bw", ""))}">'
                       f'{_forma(VERNACULO, e["et"])}</etymology>')
        for campo, nombre in (("nt", "general"), ("lem", "lema-fuente"),
                              ("st", "confianza")):
            if e.get(campo):
                out.append(f'    <note type="{nombre}">'
                           f'{_forma(ANALISIS, e[campo])}</note>')

        for i, ac in enumerate(e["acepciones"]):
            out.append(f'    <sense id="{guid}_s{i}" order="{i}">')
            if ac.get("ps"):
                out.append(f'      <grammatical-info value="{escape(ac["ps"])}"/>')
            if ac.get("ge"):
                out.append(f'      <gloss lang="{ANALISIS}">'
                           f'<text>{escape(ac["ge"])}</text></gloss>')
            if ac.get("de"):
                out.append(f'      <definition>{_forma(ANALISIS, ac["de"])}'
                           f'</definition>')
            if ac.get("sd"):
                out.append('      <trait name="semantic-domain-ddp4" '
                           f'value="{escape(ac["sd"])}"/>')
            if ac.get("xv"):
                src = f' source="{escape(ac["rf"])}"' if ac.get("rf") else ""
                out.append(f'      <example{src}>')
                out.append("        " + _forma(VERNACULO, ac["xv"]))
                if ac.get("xn"):
                    out.append(f'        <translation>'
                               f'{_forma(ANALISIS, ac["xn"])}</translation>')
                out.append("      </example>")
            if ac.get("re"):
                out.append('      <reversal type="es">'
                           f'{_forma(ANALISIS, ac["re"])}</reversal>')
            for campo, tipo in (("cf", "see-also"), ("sy", "synonym"),
                                ("an", "antonym")):
                if ac.get(campo):
                    out.append(f'      <relation type="{tipo}" '
                               f'ref="{escape(ac[campo])}"/>')
            out.append("    </sense>")

        out.append("  </entry>")

    out.append("</lift>")
    ruta.write_text("\n".join(out) + "\n", encoding="utf-8", newline="\n")


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        return 1
    origen = Path(sys.argv[1]).expanduser().resolve()
    if not origen.exists():
        print(f"No existe el archivo: {origen}")
        return 1

    try:
        filas = leer_filas(origen)
    except ErrorPlantilla as err:
        print(f"Error en la plantilla: {err}")
        return 1

    entradas, problemas = agrupar(filas)

    sfm = origen.with_suffix(".sfm")
    lift = origen.with_suffix(".lift")
    escribir_sfm(entradas, sfm)
    escribir_lift(entradas, lift)

    acepciones = sum(len(e["acepciones"]) for e in entradas)
    print(f"{len(filas)} filas -> {len(entradas)} entradas, "
          f"{acepciones} acepciones")
    print(f"  {sfm}")
    print(f"  {lift}")

    if problemas:
        print(f"\n{len(problemas)} cosas para revisar:")
        for p in problemas[:20]:
            print(f"  - {p}")
        if len(problemas) > 20:
            print(f"  ... y {len(problemas) - 20} más")
        print("\nLos archivos se generaron igual: revisa estos puntos antes de "
              "importar.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
