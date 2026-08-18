# -*- coding: utf-8 -*-
"""Salidas del diccionario: markdown, tsv, xlsx y html."""
import os, json

from . import config
from .build import sortkey, kindtag, tier, source_list, load_morphology
from .locales import Strings


def _letters(keys):
    out = {}
    for h in keys:
        k = sortkey(h)
        out.setdefault(k[0].upper() if k else "?", []).append(h)
    return out


# --------------------------------------------------------------------------
def markdown(proj, S, main, propers, glossary, counts, propmaps):
    PROP, PROP_ID = propmaps
    KIND = S.map("kind_names")
    TYPE = S.map("type_names")
    out = []
    W = out.append

    W(f"# {S('title')}"); W("")
    W(S("lang_desc")); W("")
    W(S("intro")); W("")
    W("---"); W("")
    W(f"## {S('tiers_heading')}"); W("")
    W("- " + S("tier_A", lexicon=source_list(proj)))
    W("- " + S("tier_B"))
    W("- " + S("tier_C"))
    W("")
    W(S("corpus_confirm")); W("")
    W(S("kinds_note")); W("")
    W(S("summary")); W(""); W("---"); W("")

    W(f"## {S('sec_main')}"); W("")
    letters = _letters(main)
    for L in sorted(letters):
        W(f"### {L}"); W("")
        for head in sorted(letters[L], key=sortkey):
            e = main[head]; tg = tier(e); tag = f"[{tg}]"
            if tg == "C" and e.get("corpus"):
                tag += f"·{e['corpus']['conf']}"
            conf = "  ✓korpus" if (tg in ("A", "B") and e.get("corpus")) else ""
            kt = kindtag(e["kinds"])
            ktxt = "" if kt == "palabra" else f" *({KIND.get(kt, kt)})*"
            gl = "; ".join(e["glosses"]) if e["glosses"] else "—"
            W(f"- **{head}** {tag}{ktxt} — {gl}{conf}")
        W("")

    if glossary:
        W("---"); W("")
        W(f"## {S('sec_glossary')}"); W("")
        for g in sorted(glossary, key=lambda x: sortkey(x["head"])):
            W(f"- **{g['head']}** — *{g['ind'] or '—'}*")
            if g["defn"]:
                d = g["defn"].strip()
                if len(d) > 600:
                    d = d[:600].rsplit(" ", 1)[0] + " …"
                W(f"  - {d}")
        W("")

    if propers:
        W("---"); W("")
        W(f"## {S('sec_propers')}"); W("")
        W(S("propers_note")); W("")
        plett = _letters(propers)
        for L in sorted(plett):
            W(f"### {L}"); W("")
            for h in sorted(plett[L], key=sortkey):
                pe = propers[h]
                if pe.get("src") == "corpus" and h not in PROP:
                    typ = TYPE.get("korpus", "korpus")
                else:
                    typ = TYPE.get(PROP.get(h, "lain"), "lain-lain")
                idn = PROP_ID.get(h, "") or pe.get("ind", "")
                extra = f" · {proj.gloss_name}: *{idn}*" if idn and idn.lower() != h.lower() else ""
                W(f"- **{h}** — {typ}{extra}")
            W("")

    morf = load_morphology(proj)
    if morf:
        W("---"); W("")
        W(f"## {S('sec_morphology')}"); W("")
        for row in morf:
            W(f"- **{row[0]}** — {row[1]}" + (f" · *{row[2]}*" if len(row) > 2 and row[2] else ""))
        W("")

    W("---"); W("")
    W(S("sources_footer", sources=source_list(proj)))

    path = proj.out("md")
    open(path, "w", encoding="utf-8").write("\n".join(out))
    return path


# --------------------------------------------------------------------------
def tsv(proj, S, main):
    path = proj.out("tsv")
    with open(path, "w", encoding="utf-8") as f:
        f.write("origen\tglosa\tjenis\ttingkat\tconf_corpus\tconfirmado_corpus\tlema_yunani_ibrani\n")
        for head in sorted(main, key=sortkey):
            e = main[head]; tg = tier(e)
            cc = e["corpus"]["conf"] if e.get("corpus") else ""
            conf = "si" if (tg in ("A", "B") and e.get("corpus")) else ""
            f.write(f"{head}\t{'; '.join(e['glosses'])}\t{kindtag(e['kinds'])}\t{tg}\t"
                    f"{cc}\t{conf}\t{'; '.join(e['lemmas'])}\n")
    return path


# --------------------------------------------------------------------------
def xlsx(proj, S, main, propers, glossary, counts, propmaps):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  (openpyxl no instalado: se omite el .xlsx)")
        return None

    PROP, PROP_ID = propmaps
    KIND = S.map("kind_names")
    TYPE = S.map("type_names")

    wb = Workbook()
    HFILL = PatternFill("solid", fgColor="1F4E79"); HFONT = Font(bold=True, color="FFFFFF")
    HALIGN = Alignment(horizontal="center", vertical="center")
    WRAP = Alignment(vertical="top", wrap_text=True); TOP = Alignment(vertical="top")
    A_FILL = PatternFill("solid", fgColor="E2EFDA"); B_FILL = PatternFill("solid", fgColor="FCE4D6")
    C_FILL = PatternFill("solid", fgColor="DDEBF7")
    thin = Side(style="thin", color="D9D9D9")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr(ws, n):
        for c in range(1, n + 1):
            cell = ws.cell(1, c)
            cell.fill = HFILL; cell.font = HFONT; cell.alignment = HALIGN; cell.border = BORDER
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(n)}{ws.max_row}"

    ws0 = wb.active; ws0.title = S("readme_sheet")
    for i, line in enumerate(S.list("readme_lines", sources=source_list(proj)), 1):
        ws0.cell(i, 1, line).font = Font(bold=(i == 1), size=13 if i == 1 else 11)
    ws0.column_dimensions["A"].width = 100
    ws0.sheet_view.showGridLines = False

    ws1 = wb.create_sheet(S("tab_main"))
    ws1.append([S("col_source"), S("col_gloss"), S("col_kind"), S("col_tier"),
                S("col_corpus"), S("col_lemma")])
    for head in sorted(main, key=sortkey):
        e = main[head]; tg = tier(e)
        cc = e["corpus"]["conf"] if e.get("corpus") else ""
        if tg in ("A", "B") and cc:
            cc = "✓ " + cc
        ws1.append([head, "; ".join(e["glosses"]), KIND.get(kindtag(e["kinds"]), ""),
                    tg, cc, "; ".join(e["lemmas"])])
    for r in range(2, ws1.max_row + 1):
        tv = ws1.cell(r, 4).value
        fill = A_FILL if tv == "A" else (B_FILL if tv == "B" else C_FILL)
        for c in range(1, 7):
            cell = ws1.cell(r, c)
            cell.border = BORDER
            cell.alignment = WRAP if c == 2 else TOP
            if c in (1, 4):
                cell.fill = fill
        ws1.cell(r, 1).font = Font(bold=True)
        ws1.cell(r, 4).alignment = HALIGN
    for col, w in zip("ABCDEF", (28, 50, 9, 8, 13, 26)):
        ws1.column_dimensions[col].width = w
    hdr(ws1, 6)

    if glossary:
        ws2 = wb.create_sheet(S("tab_glossary"))
        ws2.append([S("col_source"), S("col_gloss"), S("col_definition")])
        for g in sorted(glossary, key=lambda x: sortkey(x["head"])):
            ws2.append([g["head"], g["ind"], g["defn"]])
        for r in range(2, ws2.max_row + 1):
            for c in range(1, 4):
                cell = ws2.cell(r, c)
                cell.border = BORDER
                cell.alignment = WRAP if c == 3 else TOP
            ws2.cell(r, 1).font = Font(bold=True)
            ws2.cell(r, 2).font = Font(italic=True)
        for col, w in zip("ABC", (32, 26, 100)):
            ws2.column_dimensions[col].width = w
        hdr(ws2, 3)

    if propers:
        ws3 = wb.create_sheet(S("tab_propers"))
        ws3.append([S("col_source"), S("col_type"), S("col_gloss"), S("col_srcname")])
        for head in sorted(propers, key=sortkey):
            pe = propers[head]
            if pe.get("src") == "corpus" and head not in PROP:
                typ = TYPE.get("korpus", "korpus")
                srcp = f"korpus {proj.bt['name']}" if proj.bt else "korpus"
            else:
                typ = TYPE.get(PROP.get(head, "lain"), "lain-lain")
                srcp = "Alkitab"
            idn = PROP_ID.get(head, "") or pe.get("ind", "") or head
            ws3.append([head, typ, idn, srcp])
        for r in range(2, ws3.max_row + 1):
            for c in range(1, 5):
                cell = ws3.cell(r, c); cell.border = BORDER; cell.alignment = TOP
            ws3.cell(r, 1).font = Font(bold=True)
        for col, w in zip("ABCD", (30, 14, 30, 26)):
            ws3.column_dimensions[col].width = w
        hdr(ws3, 4)

    morf = load_morphology(proj)
    if morf:
        ws4 = wb.create_sheet(S("tab_morphology"))
        ws4.append([S("col_affix"), S("col_function"), S("col_example")])
        for row in morf:
            ws4.append(list(row) + [""] * (3 - len(row)))
        for c in range(1, 4):
            cell = ws4.cell(1, c)
            cell.fill = HFILL; cell.font = HFONT; cell.alignment = HALIGN; cell.border = BORDER
        for r in range(2, ws4.max_row + 1):
            for c in range(1, 4):
                cell = ws4.cell(r, c); cell.border = BORDER; cell.alignment = TOP
            ws4.cell(r, 1).font = Font(bold=True)
        ws4.freeze_panes = "A2"
        for col, w in zip("ABC", (24, 42, 42)):
            ws4.column_dimensions[col].width = w

    path = proj.out("xlsx")
    try:
        wb.save(path)
        print("  xlsx OK")
    except PermissionError:
        path = proj.out("xlsx").replace(".xlsx", "_v2.xlsx")
        wb.save(path)
        print("  xlsx BLOQUEADO (¿abierto en Excel?) -> guardado como", os.path.basename(path))
    return path


# --------------------------------------------------------------------------
def html(proj, S, main, propers, glossary, counts, propmaps):
    PROP, PROP_ID = propmaps

    def hmeta(e):
        parts = []
        if e["lemmas"]:
            parts.append("← " + "; ".join(e["lemmas"]))
        if e.get("corpus"):
            c = e["corpus"]
            parts.append(("✓korpus " if tier(e) in ("A", "B") else "korpus ")
                         + f"p{c['p']} n{c['n']}")
        return " · ".join(parts)

    main_rows = [{"t": h, "i": "; ".join(main[h]["glosses"]), "lv": tier(main[h]),
                  "cf": (main[h]["corpus"]["conf"]
                         if (tier(main[h]) == "C" and main[h].get("corpus")) else ""),
                  "k": kindtag(main[h]["kinds"]), "m": hmeta(main[h])}
                 for h in sorted(main, key=sortkey)]
    glos_rows = [{"t": g["head"], "i": g["ind"], "d": g["defn"]}
                 for g in sorted(glossary, key=lambda x: sortkey(x["head"]))]

    def prow(h):
        pe = propers[h]
        corpusname = pe.get("src") == "corpus" and h not in PROP
        return {"t": h,
                "ty": "korpus" if corpusname else PROP.get(h, "lain"),
                "id": (PROP_ID.get(h, "") or pe.get("ind", "")) if
                      (PROP_ID.get(h, "") or pe.get("ind", "")).lower() != h.lower() else ""}

    prop_rows = [prow(h) for h in sorted(propers, key=sortkey)]
    morf_rows = load_morphology(proj)

    DATA = {"main": main_rows, "glos": glos_rows, "prop": prop_rows, "morf": morf_rows,
            "counts": {"main": len(main_rows), "A": counts["A"], "B": counts["B"],
                       "glos": len(glos_rows), "prop": len(prop_rows)}}

    tpl = open(os.path.join(config.TEMPLATES, "kamus.html"), encoding="utf-8").read()
    subs = {
        "__DATA__": json.dumps(DATA, ensure_ascii=False),
        "__UI__": json.dumps(S.ui(), ensure_ascii=False),
        "__GLOSS_ISO__": proj.gloss_iso or "id",
        "__TITLE__": S("title"),
        "__SUBTITLE__": S("lang_desc").replace("**", ""),
        "__SEARCH__": S("search_ph"),
        "__CHIP_A__": S("chip_A"), "__CHIP_B__": S("chip_B"), "__CHIP_C__": S("chip_C"),
        "__KIND_W__": S.map("kind_names").get("palabra", "kata"),
        "__KIND_P__": S.map("kind_names").get("frase", "frasa"),
        "__KIND_R__": S.map("kind_names").get("raiz", "akar"),
        "__TAB_MAIN__": S("tab_main"), "__TAB_GLOS__": S("tab_glossary"),
        "__TAB_PROP__": S("tab_propers"), "__TAB_MORF__": S("tab_morphology"),
        "__LINK_INTER__": S("link_to_inter"),
        "__INTER_FILE__": proj.inter_basename + ".html",
    }
    for k, v in subs.items():
        tpl = tpl.replace(k, v)

    path = proj.out("html")
    open(path, "w", encoding="utf-8").write(tpl)
    return path


# --------------------------------------------------------------------------
def dump(proj, main, propers, propmaps):
    """Volcado que consume el interlineal."""
    PROP, PROP_ID = propmaps
    d = {"main": {h: {"g": main[h]["glosses"], "lv": tier(main[h])} for h in main},
         "prop": {h: (PROP_ID.get(h, "") or propers[h].get("ind", "") or h) for h in propers}}
    path = proj.workfile("dict_dump.json")
    json.dump(d, open(path, "w", encoding="utf-8"), ensure_ascii=False)
    return path
